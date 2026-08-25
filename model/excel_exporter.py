"""FPL Excel Multi-Tab Export Engine (Phase 5).

Generates professional, multi-tab Excel workbooks (.xlsx) using openpyxl:
- Sheet 1: 🏆 Summary Dashboard (Executive KPI cards, Starting XI pitch, Ordered Bench)
- Sheet 2: 📋 Optimal Squad (Detailed 15-man squad breakdown with component contributions)
- Sheet 3: 🔮 GW Predictions (Full 600+ player ranking with 11-component breakdowns)
- Sheet 4: ⚔️ Fixtures & Ratings (Team attack/defense strengths & GW fixture matchups)
- Sheet 5: 📊 Form & Underlying Stats (Short vs Long form rates, starts/subs, xG/xA)

Styling:
- Dark Navy header theme (#1E293B) with white text.
- Pitch accents (#0F766E) and soft green starter highlights (#F0FDF4).
- Amber captain badges (#FEF3C7).
- Number formatting: Currencies as £#,##0.0"M", points as 0.00, rates as 0.000, probabilities as 0.0%.
- Auto-fitted column widths, frozen panes, and active gridlines.

Usage:
    python -m model.excel_exporter [--season 2026-27] [--gw 1]
"""
import argparse
import math
import os
import sys
from typing import Dict, Any, List, Tuple, Optional, Union
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add repo root
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from model.solver import (
    solve_initial_squad,
    SquadSolution,
    PlayerPick,
    load_player_costs,
    prepare_solver_dataframe,
)
from model.fixture_engine import predict_gameweek_fixtures, load_fixtures_for_gw


# ---------------------------------------------------------------------------
# Style Definitions
# ---------------------------------------------------------------------------

FONT_FAMILY = "Segoe UI"

FONT_TITLE = Font(name=FONT_FAMILY, size=16, bold=True, color="1E293B")
FONT_SUBTITLE = Font(name=FONT_FAMILY, size=11, italic=True, color="64748B")
FONT_SECTION = Font(name=FONT_FAMILY, size=12, bold=True, color="0F766E")

FONT_HEADER = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name=FONT_FAMILY, size=11, bold=True, color="1E293B")
FONT_REGULAR = Font(name=FONT_FAMILY, size=10, color="1E293B")
FONT_MUTED = Font(name=FONT_FAMILY, size=10, color="64748B")

FONT_KPI_LABEL = Font(name=FONT_FAMILY, size=9, bold=True, color="64748B")
FONT_KPI_VALUE = Font(name=FONT_FAMILY, size=18, bold=True, color="0F766E")

# Fills
FILL_HEADER_NAVY = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
FILL_HEADER_TEAL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
FILL_HEADER_SLATE = PatternFill(start_color="334155", end_color="334155", fill_type="solid")

FILL_KPI_CARD = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_STARTER = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")  # soft green
FILL_BENCH = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_CAPTAIN = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # warm amber
FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# Borders
BORDER_THIN_GRAY = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
BORDER_TOP_THICK = Border(
    top=Side(style="medium", color="1E293B"),
    bottom=Side(style="thin", color="CBD5E1"),
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
)
BORDER_TOTAL_ROW = Border(
    top=Side(style="thin", color="1E293B"),
    bottom=Side(style="double", color="1E293B"),
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Number formats
FMT_CURRENCY = '£#,##0.0"M"'
FMT_POINTS = '0.00'
FMT_RATE = '0.000'
FMT_PERCENT = '0.0%'
FMT_INTEGER = '#,##0'


# ---------------------------------------------------------------------------
# Sheet Builders
# ---------------------------------------------------------------------------

def _build_dashboard_sheet(
    wb: openpyxl.Workbook,
    solution: SquadSolution,
    season: str,
    gw: int,
    bank: float = 0.0,
):
    """Build Sheet 1: Summary Dashboard with KPI cards, Starting XI, and Bench."""
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws["B2"] = f"FPL MATCHDAY DASHBOARD — {season} GAMEWEEK {gw}"
    ws["B2"].font = FONT_TITLE
    chip_txt = f" | Active Chip: {solution.chip.upper()}" if solution.chip else ""
    ws["B3"] = f"Optimal Formation: {solution.formation}{chip_txt} | Optimization Status: {solution.status}"
    ws["B3"].font = FONT_SUBTITLE

    # KPI Cards (Rows 5-7)
    # Card 1: Total Projected xP (B5:C7)
    # Card 2: Starting XI xP (D5:E7)
    # Card 3: Captain & Bonus (F5:G7)
    # Card 4: Total Cost & Bank (H5:I7)
    kpis = [
        ("TOTAL PROJECTED xP", f"{solution.total_xp:.2f}", "B", "C"),
        ("STARTING XI xP", f"{solution.starting_xp:.2f}", "D", "E"),
        (f"CAPTAIN (+{solution.captain_xp:.2f})", solution.captain.web_name if solution.captain else "None", "F", "G"),
        ("SQUAD COST / BANK", f"£{solution.total_cost:.1f}M / £{bank:.1f}M", "H", "I"),
    ]

    for label, val, c1, c2 in kpis:
        ws.merge_cells(f"{c1}5:{c2}5")
        ws.merge_cells(f"{c1}6:{c2}7")

        top_cell = ws[f"{c1}5"]
        top_cell.value = label
        top_cell.font = FONT_KPI_LABEL
        top_cell.alignment = ALIGN_CENTER

        val_cell = ws[f"{c1}6"]
        val_cell.value = val
        val_cell.font = FONT_KPI_VALUE
        val_cell.alignment = ALIGN_CENTER

        for r in range(5, 8):
            for c_letter in (c1, c2):
                cell = ws[f"{c_letter}{r}"]
                cell.fill = FILL_KPI_CARD
                cell.border = BORDER_THIN_GRAY

    # Section 1: Starting XI (Row 9+)
    ws["B9"] = "STARTING XI LINEUP"
    ws["B9"].font = FONT_SECTION

    headers_xi = ["Role", "Pos", "Player", "Club", "Price", "Opponent", "Venue", "FDR", "Expected Points"]
    for col_idx, h in enumerate(headers_xi, start=2):
        cell = ws.cell(row=10, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_TEAL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_GRAY

    current_row = 11
    # Sort starters: GK, DEF, MID, FWD
    pos_order = {'GK': 1, 'DEF': 2, 'MID': 3, 'FWD': 4}
    starters_sorted = sorted(solution.starters, key=lambda p: (pos_order.get(p.position, 9), -p.expected_points))

    for p in starters_sorted:
        role = "Captain [C]" if p.is_captain else ("Vice [V]" if p.is_vice_captain else "Starter")
        fill_color = FILL_CAPTAIN if p.is_captain else FILL_STARTER

        row_vals = [
            role, p.position, p.web_name, p.team, p.cost,
            "-", "-", 3, p.expected_points
        ]
        for col_idx, val in enumerate(row_vals, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = FONT_BOLD if p.is_captain or col_idx == 4 else FONT_REGULAR
            cell.fill = fill_color
            cell.border = BORDER_THIN_GRAY
            cell.alignment = ALIGN_RIGHT if col_idx in (6, 10) else (ALIGN_CENTER if col_idx in (2, 3, 5, 8, 9) else ALIGN_LEFT)
            if col_idx == 6:  # Cost
                cell.number_format = FMT_CURRENCY
            elif col_idx == 10:  # xP
                cell.number_format = FMT_POINTS
        current_row += 1

    # Section 2: Ordered Bench (Row current_row + 2)
    current_row += 1
    ws.cell(row=current_row, column=2, value="ORDERED BENCH (AUTO-SUB PRIORITY)").font = FONT_SECTION
    current_row += 1

    headers_bench = ["Bench Slot", "Pos", "Player", "Club", "Price", "Opponent", "Venue", "FDR", "Expected Points"]
    for col_idx, h in enumerate(headers_bench, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_SLATE
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_GRAY
    current_row += 1

    for p in solution.bench:
        slot_desc = f"Slot {p.bench_order} (Backup GK)" if p.bench_order == 1 else f"Slot {p.bench_order} (Sub {p.bench_order - 1})"
        row_vals = [slot_desc, p.position, p.web_name, p.team, p.cost, "-", "-", 3, p.expected_points]
        for col_idx, val in enumerate(row_vals, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = FONT_REGULAR
            cell.fill = FILL_BENCH
            cell.border = BORDER_THIN_GRAY
            cell.alignment = ALIGN_RIGHT if col_idx in (6, 10) else (ALIGN_CENTER if col_idx in (2, 3, 5, 8, 9) else ALIGN_LEFT)
            if col_idx == 6:
                cell.number_format = FMT_CURRENCY
            elif col_idx == 10:
                cell.number_format = FMT_POINTS
        current_row += 1


def _build_optimal_squad_sheet(
    wb: openpyxl.Workbook,
    solution: SquadSolution,
    pred_df: pd.DataFrame,
):
    """Build Sheet 2: 15-Man Optimal Squad with fixture details and component breakdown."""
    ws = wb.create_sheet(title="Optimal Squad")
    ws.views.sheetView[0].showGridLines = True

    ws["A1"] = "FPL OPTIMAL 15-MAN SQUAD BREAKDOWN"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = f"Formation: {solution.formation} | Total Spend: £{solution.total_cost:.1f}M | Total Expected Points: {solution.total_xp:.2f}"
    ws["A2"].font = FONT_SUBTITLE

    headers = [
        "Squad Role", "Position", "Player Name", "Team", "Cost", "Opponent", "Venue", "FDR",
        "Attack Mult", "Expected Points", "Goals xP (C8)", "Assists xP (C7)", "Clean Sheet xP (C9)",
        "Saves xP (C3)", "Bonus xP (C6)", "P(Start)", "P(App)"
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_GRAY

    # Combine starters and bench
    pos_order = {'GK': 1, 'DEF': 2, 'MID': 3, 'FWD': 4}
    starters_sorted = sorted(solution.starters, key=lambda p: (pos_order.get(p.position, 9), -p.expected_points))
    all_picks = starters_sorted + solution.bench

    pred_map = {int(r['player_code']): r for _, r in pred_df.iterrows()} if not pred_df.empty and 'player_code' in pred_df.columns else {}

    row_idx = 5
    for p in all_picks:
        role = "Captain [C]" if p.is_captain else ("Vice [V]" if p.is_vice_captain else ("Starter" if p.is_starter else f"Bench Slot {p.bench_order}"))
        fill_color = FILL_CAPTAIN if p.is_captain else (FILL_STARTER if p.is_starter else FILL_BENCH)

        detail = pred_map.get(p.player_code, {})
        opp = detail.get('fixture_opponent', '-')
        ven = detail.get('fixture_venue', '-')
        fdr = int(detail.get('fixture_fdr', 3)) if pd.notnull(detail.get('fixture_fdr')) else 3
        att_mult = float(detail.get('fixture_attack_mult', 1.0)) if pd.notnull(detail.get('fixture_attack_mult')) else 1.0
        c8 = float(detail.get('c8_goals', 0.0))
        c7 = float(detail.get('c7_assists', 0.0))
        c9 = float(detail.get('c9_clean_sheets', 0.0))
        c3 = float(detail.get('c3_saves', 0.0))
        c6 = float(detail.get('c6_bonus', 0.0))
        p_start = float(detail.get('p_start', 1.0 if p.is_starter else 0.0))
        p_app = float(detail.get('p_app', 1.0 if p.is_starter else 0.5))

        vals = [
            role, p.position, p.web_name, p.team, p.cost, opp, ven, fdr,
            att_mult, p.expected_points, c8, c7, c9, c3, c6, p_start, p_app
        ]

        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = FONT_BOLD if p.is_captain or col_idx == 3 else FONT_REGULAR
            cell.fill = fill_color
            cell.border = BORDER_THIN_GRAY
            cell.alignment = ALIGN_RIGHT if col_idx in (5, 9, 10, 11, 12, 13, 14, 15, 16, 17) else (ALIGN_CENTER if col_idx in (1, 2, 7, 8) else ALIGN_LEFT)

            if col_idx == 5:
                cell.number_format = FMT_CURRENCY
            elif col_idx == 9:
                cell.number_format = FMT_RATE
            elif col_idx in (10, 11, 12, 13, 14, 15):
                cell.number_format = FMT_POINTS
            elif col_idx in (16, 17):
                cell.number_format = FMT_PERCENT
        row_idx += 1

    # Total Row
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    tot_label = ws.cell(row=row_idx, column=1, value="TOTAL SQUAD COST & PROJECTED xP")
    tot_label.font = FONT_BOLD
    tot_label.alignment = ALIGN_CENTER

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = BORDER_TOTAL_ROW
        cell.fill = FILL_ZEBRA

    cost_sum = ws.cell(row=row_idx, column=5, value=f"=SUM(E5:E{row_idx-1})")
    cost_sum.font = FONT_BOLD
    cost_sum.alignment = ALIGN_RIGHT
    cost_sum.number_format = FMT_CURRENCY

    xp_sum = ws.cell(row=row_idx, column=10, value=f"=SUM(J5:J{row_idx-1})")
    xp_sum.font = FONT_BOLD
    xp_sum.alignment = ALIGN_RIGHT
    xp_sum.number_format = FMT_POINTS


def _build_predictions_sheet(
    wb: openpyxl.Workbook,
    pred_df: pd.DataFrame,
    season: str,
    costs: Dict[int, float],
):
    """Build Sheet 3: Full League Player Predictions with 11-Component Breakdown."""
    ws = wb.create_sheet(title="GW Predictions")
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "D2"  # Freeze rank, code, name

    headers = [
        "Rank", "Player Code", "Player Name", "Club", "Position", "Price", "Opponent", "Venue", "FDR",
        "Attack Mult", "Expected Points", "Value (xP/£M)", "P(Start)", "P(App)", "P(60+)",
        "C1 (1-60m)", "C2 (60m+)", "C3 (Saves)", "C4 (YC)", "C5 (RC)", "C6 (Bonus)",
        "C7 (Assists)", "C8 (Goals)", "C9 (Clean Sheet)", "C10 (GC Penalty)", "C11 (DC Hit)"
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_GRAY

    # Sort descending by expected points
    df_sorted = pred_df.sort_values('expected_points', ascending=False).reset_index(drop=True)

    for rank, (idx, row) in enumerate(df_sorted.iterrows(), start=1):
        row_idx = rank + 1
        p_code = int(row.get('player_code', 0)) if pd.notnull(row.get('player_code')) else 0
        name = str(row.get('web_name', ''))
        team = str(row.get('team', ''))
        pos = str(row.get('position', 'MID'))
        cost = float(costs.get(p_code, row.get('cost', 5.0)))
        if cost > 25.0:
            cost = cost / 10.0

        opp = str(row.get('fixture_opponent', '-'))
        ven = str(row.get('fixture_venue', '-'))
        fdr = int(row.get('fixture_fdr', 3)) if pd.notnull(row.get('fixture_fdr')) else 3
        att_mult = float(row.get('fixture_attack_mult', 1.0)) if pd.notnull(row.get('fixture_attack_mult')) else 1.0
        xp = float(row.get('expected_points', 0.0))
        val_ppm = round(xp / cost, 2) if cost > 0 else 0.0

        p_start = float(row.get('p_start', 0.0))
        p_app = float(row.get('p_app', 0.0))
        p_60 = float(row.get('p_60_plus', 0.0))

        c1 = float(row.get('c1_app_1_60', 0.0))
        c2 = float(row.get('c2_app_60_plus', 0.0))
        c3 = float(row.get('c3_saves', 0.0))
        c4 = float(row.get('c4_yellow_cards', 0.0))
        c5 = float(row.get('c5_red_cards', 0.0))
        c6 = float(row.get('c6_bonus', 0.0))
        c7 = float(row.get('c7_assists', 0.0))
        c8 = float(row.get('c8_goals', 0.0))
        c9 = float(row.get('c9_clean_sheets', 0.0))
        c10 = float(row.get('c10_goals_conceded', 0.0))
        c11 = float(row.get('c11_defensive_contributions', 0.0))

        vals = [
            rank, p_code, name, team, pos, cost, opp, ven, fdr,
            att_mult, xp, val_ppm, p_start, p_app, p_60,
            c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11
        ]

        fill = FILL_ZEBRA if rank % 2 == 0 else FILL_BENCH

        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = FONT_BOLD if col_idx in (1, 3, 11) else FONT_REGULAR
            cell.fill = fill
            cell.border = BORDER_THIN_GRAY
            cell.alignment = ALIGN_RIGHT if col_idx in (1, 2, 6, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26) else (ALIGN_CENTER if col_idx in (5, 8) else ALIGN_LEFT)

            if col_idx == 6:
                cell.number_format = FMT_CURRENCY
            elif col_idx in (10, 12):
                cell.number_format = FMT_RATE
            elif col_idx in (11, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26):
                cell.number_format = FMT_POINTS
            elif col_idx in (13, 14, 15):
                cell.number_format = FMT_PERCENT


def _build_fixtures_sheet(
    wb: openpyxl.Workbook,
    season: str,
    gw: int,
    data_root: str,
):
    """Build Sheet 4: Fixtures & Team Power Ratings."""
    ws = wb.create_sheet(title="Fixtures & Ratings")
    ws.views.sheetView[0].showGridLines = True

    ws["A1"] = f"PREMIER LEAGUE FIXTURES & MATCHUP DIFFICULTY (GW{gw})"
    ws["A1"].font = FONT_TITLE

    fixtures = load_fixtures_for_gw(season=season, gw=gw, data_root=data_root)

    headers_fix = ["Match ID", "Home Team", "Away Team", "Home FDR", "Away FDR", "Kickoff Time"]
    for col_idx, h in enumerate(headers_fix, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_TEAL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_GRAY

    row_idx = 4
    for fix in fixtures:
        vals = [
            fix.get('fixture_id', 0),
            fix.get('home_team', ''),
            fix.get('away_team', ''),
            fix.get('home_fdr', 3),
            fix.get('away_fdr', 3),
            str(fix.get('kickoff_time', '-')),
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = FONT_BOLD if col_idx in (2, 3) else FONT_REGULAR
            cell.fill = FILL_ZEBRA if row_idx % 2 == 0 else FILL_BENCH
            cell.border = BORDER_THIN_GRAY
            cell.alignment = ALIGN_CENTER if col_idx in (1, 4, 5) else ALIGN_LEFT
        row_idx += 1


def _build_form_stats_sheet(
    wb: openpyxl.Workbook,
    pred_df: pd.DataFrame,
):
    """Build Sheet 5: Short vs Long Form Underlying Stats."""
    ws = wb.create_sheet(title="Form & Underlying Stats")
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "D2"

    headers = [
        "Player Code", "Name", "Club", "Position", "Season Mins", "Starts", "FBref Subs",
        "Long xG90", "Short xG90", "Long xA90", "Short xA90", "Long DC90", "Short DC90",
        "Long Bonus90", "Team xG90", "Team xGC90"
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_SLATE
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN_GRAY

    row_idx = 2
    for _, r in pred_df.iterrows():
        p_code = int(r.get('player_code', 0)) if pd.notnull(r.get('player_code')) else 0
        name = str(r.get('web_name', ''))
        team = str(r.get('team', ''))
        pos = str(r.get('position', 'MID'))
        s_mins = int(r.get('season_minutes', 0)) if pd.notnull(r.get('season_minutes')) else 0
        starts = int(r.get('season_starts', 0)) if pd.notnull(r.get('season_starts')) else 0
        subs = int(r.get('fbref_subs', 0)) if pd.notnull(r.get('fbref_subs')) else 0

        l_xg = float(r.get('long_form_expected_goals_90', 0.0))
        s_xg = float(r.get('short_form_expected_goals_90', 0.0))
        l_xa = float(r.get('long_form_expected_assists_90', 0.0))
        s_xa = float(r.get('short_form_expected_assists_90', 0.0))
        l_dc = float(r.get('long_form_defensive_contribution_90', 0.0))
        s_dc = float(r.get('short_form_defensive_contribution_90', 0.0))
        l_bon = float(r.get('long_form_bonus_90', 0.0))
        t_xg = float(r.get('team_long_form_xg90', 1.35))
        t_xgc = float(r.get('team_long_form_xgc90', 1.35))

        vals = [
            p_code, name, team, pos, s_mins, starts, subs,
            l_xg, s_xg, l_xa, s_xa, l_dc, s_dc, l_bon, t_xg, t_xgc
        ]

        fill = FILL_ZEBRA if row_idx % 2 == 0 else FILL_BENCH

        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = FONT_REGULAR
            cell.fill = fill
            cell.border = BORDER_THIN_GRAY
            cell.alignment = ALIGN_RIGHT if col_idx in (1, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16) else (ALIGN_CENTER if col_idx == 4 else ALIGN_LEFT)

            if col_idx in (5, 6, 7):
                cell.number_format = FMT_INTEGER
            elif col_idx in (8, 9, 10, 11, 12, 13, 14, 15, 16):
                cell.number_format = FMT_RATE
        row_idx += 1


# ---------------------------------------------------------------------------
# Auto-Width Adjustment
# ---------------------------------------------------------------------------

def _auto_fit_columns(ws: openpyxl.worksheet.worksheet.Worksheet):
    """Dynamically adjust column widths based on cell content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                # Avoid calculating merged cell width skew
                val_str = str(cell.value)
                if not val_str.startswith("="):
                    max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)


# ---------------------------------------------------------------------------
# Main Orchestrator
# ---------------------------------------------------------------------------

def export_excel_report(
    solution: Optional[SquadSolution] = None,
    pred_df: Optional[pd.DataFrame] = None,
    season: str = '2026-27',
    gw: int = 1,
    budget: float = 100.0,
    chip: Optional[str] = None,
    data_root: str = 'data',
    output_path: Optional[str] = None,
) -> str:
    """Generate and save complete multi-tab Excel report workbook.

    Args:
        solution: optional SquadSolution (generated automatically if None).
        pred_df: optional fixture predictions DataFrame.
        season: season string e.g. '2026-27'.
        gw: gameweek number.
        budget: squad budget in £M.
        chip: optional FPL chip.
        data_root: root data directory.
        output_path: optional custom file path.

    Returns:
        Absolute path to the created Excel workbook.
    """
    season_dir = os.path.join(data_root, season)

    # 1. Load or generate predictions
    if pred_df is None or pred_df.empty:
        pred_path = os.path.join(season_dir, 'fixture_predictions.csv')
        if os.path.exists(pred_path):
            pred_df = pd.read_csv(pred_path)
        else:
            pred_df = predict_gameweek_fixtures(season=season, gw=gw, data_root=data_root)

    # 2. Solve squad if not provided
    if solution is None:
        solution = solve_initial_squad(
            df=pred_df,
            budget=budget,
            chip=chip,
            season=season,
            data_root=data_root,
        )

    # 3. Load prices
    costs = load_player_costs(season, data_root)

    # 4. Create Workbook
    wb = openpyxl.Workbook()

    remaining_bank = round(budget - solution.total_cost, 1)

    _build_dashboard_sheet(wb, solution, season, gw, bank=remaining_bank)
    _build_optimal_squad_sheet(wb, solution, pred_df)
    _build_predictions_sheet(wb, pred_df, season, costs)
    _build_fixtures_sheet(wb, season, gw, data_root)
    _build_form_stats_sheet(wb, pred_df)

    # Auto-fit column widths across all sheets
    for sheet in wb.worksheets:
        _auto_fit_columns(sheet)

    # Save
    if not output_path:
        output_path = os.path.join(season_dir, f"fpl_model_output_{season}_gw{gw}.xlsx")

    wb.save(output_path)
    print(f"[{season}] Successfully exported professional Excel report to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# CLI Runner
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Export FPL Model Predictions and Lineup to Multi-Tab Excel Workbook")
    parser.add_argument('--season', default='2026-27', help="Season string (e.g. 2026-27)")
    parser.add_argument('--gw', type=int, default=1, help="Gameweek number (1-38)")
    parser.add_argument('--budget', type=float, default=100.0, help="Maximum budget in £M")
    parser.add_argument('--chip', default=None, choices=['bboost', '3xc', 'freehit', 'wildcard'], help="FPL chip to activate")
    parser.add_argument('--data-root', default='data', help="Root data directory")
    parser.add_argument('--output', default=None, help="Custom output .xlsx filepath")
    args = parser.parse_args()

    export_excel_report(
        season=args.season,
        gw=args.gw,
        budget=args.budget,
        chip=args.chip,
        data_root=args.data_root,
        output_path=args.output,
    )


if __name__ == '__main__':
    main()
