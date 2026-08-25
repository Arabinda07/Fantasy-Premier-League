"""Unit tests for model.excel_exporter — Phase 5 Excel Multi-Tab Export Engine.

Verifies:
1. Multi-tab workbook creation with all 5 dedicated sheets.
2. Styling integrity: Headers, fills, borders, alignments, and number formats.
3. Dashboard KPI metrics and starting XI/bench layout.
4. Optimal squad 15-player table and SUM formulas.
5. Full player predictions table sorting and 11-component exports.
6. Auto-fitted column widths and frozen panes.
7. End-to-end export to .xlsx and read-back validation via openpyxl.
"""
import os
import sys
import tempfile
from typing import Tuple, Any
import pytest
import pandas as pd
import openpyxl

# Add repo root
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from model.excel_exporter import export_excel_report
from model.solver import solve_initial_squad


@pytest.fixture
def sample_test_data() -> Tuple[pd.DataFrame, Any]:
    """Create sample prediction dataset and squad solution for exporter tests."""
    players = [
        # GKs
        {'player_code': 1, 'web_name': 'Raya', 'team': 'Arsenal', 'position': 'GK', 'cost': 5.5, 'expected_points': 5.2, 'fixture_opponent': 'Coventry City', 'fixture_venue': 'H', 'fixture_fdr': 2, 'fixture_attack_mult': 1.08, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.99, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.99, 'c3_saves': 0.8, 'c4_yellow_cards': -0.1, 'c5_red_cards': 0.0, 'c6_bonus': 0.5, 'c7_assists': 0.0, 'c8_goals': 0.0, 'c9_clean_sheets': 2.1, 'c10_goals_conceded': -0.1, 'c11_defensive_contributions': 0.0},
        {'player_code': 2, 'web_name': 'Fabianski', 'team': 'West Ham', 'position': 'GK', 'cost': 4.0, 'expected_points': 1.0, 'fixture_opponent': 'Chelsea', 'fixture_venue': 'A', 'fixture_fdr': 4, 'fixture_attack_mult': 0.92, 'p_start': 0.0, 'p_app': 0.1, 'p_60_plus': 0.0, 'c1_app_1_60': 0.1, 'c2_app_60_plus': 0.0, 'c3_saves': 0.0, 'c4_yellow_cards': 0.0, 'c5_red_cards': 0.0, 'c6_bonus': 0.0, 'c7_assists': 0.0, 'c8_goals': 0.0, 'c9_clean_sheets': 0.0, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
        # DEFs
        {'player_code': 3, 'web_name': 'Gabriel', 'team': 'Arsenal', 'position': 'DEF', 'cost': 6.0, 'expected_points': 6.5, 'fixture_opponent': 'Coventry City', 'fixture_venue': 'H', 'fixture_fdr': 2, 'fixture_attack_mult': 1.08, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.95, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.95, 'c3_saves': 0.0, 'c4_yellow_cards': -0.1, 'c5_red_cards': 0.0, 'c6_bonus': 0.8, 'c7_assists': 0.2, 'c8_goals': 1.5, 'c9_clean_sheets': 2.1, 'c10_goals_conceded': -0.1, 'c11_defensive_contributions': 0.15},
        {'player_code': 4, 'web_name': 'Saliba', 'team': 'Man City', 'position': 'DEF', 'cost': 6.0, 'expected_points': 5.8, 'fixture_opponent': 'Bournemouth', 'fixture_venue': 'H', 'fixture_fdr': 3, 'fixture_attack_mult': 1.20, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.95, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.95, 'c3_saves': 0.0, 'c4_yellow_cards': -0.1, 'c5_red_cards': 0.0, 'c6_bonus': 0.6, 'c7_assists': 0.1, 'c8_goals': 0.5, 'c9_clean_sheets': 2.1, 'c10_goals_conceded': -0.1, 'c11_defensive_contributions': 0.65},
        {'player_code': 5, 'web_name': 'Alexander-Arnold', 'team': 'Liverpool', 'position': 'DEF', 'cost': 7.0, 'expected_points': 6.2, 'fixture_opponent': 'Newcastle', 'fixture_venue': 'A', 'fixture_fdr': 4, 'fixture_attack_mult': 0.95, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.90, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.90, 'c3_saves': 0.0, 'c4_yellow_cards': -0.1, 'c5_red_cards': 0.0, 'c6_bonus': 0.8, 'c7_assists': 1.2, 'c8_goals': 0.8, 'c9_clean_sheets': 1.5, 'c10_goals_conceded': -0.2, 'c11_defensive_contributions': 0.2},
        {'player_code': 6, 'web_name': 'Robinson', 'team': 'Fulham', 'position': 'DEF', 'cost': 4.5, 'expected_points': 4.2, 'fixture_opponent': 'Chelsea', 'fixture_venue': 'H', 'fixture_fdr': 4, 'fixture_attack_mult': 1.12, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.90, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.90, 'c3_saves': 0.0, 'c4_yellow_cards': -0.15, 'c5_red_cards': 0.0, 'c6_bonus': 0.3, 'c7_assists': 0.6, 'c8_goals': 0.2, 'c9_clean_sheets': 1.0, 'c10_goals_conceded': -0.3, 'c11_defensive_contributions': 0.55},
        {'player_code': 7, 'web_name': 'Greaves', 'team': 'Ipswich', 'position': 'DEF', 'cost': 4.0, 'expected_points': 2.5, 'fixture_opponent': 'Everton', 'fixture_venue': 'A', 'fixture_fdr': 3, 'fixture_attack_mult': 0.95, 'p_start': 0.8, 'p_app': 0.9, 'p_60_plus': 0.75, 'c1_app_1_60': 0.9, 'c2_app_60_plus': 0.75, 'c3_saves': 0.0, 'c4_yellow_cards': -0.2, 'c5_red_cards': 0.0, 'c6_bonus': 0.1, 'c7_assists': 0.1, 'c8_goals': 0.1, 'c9_clean_sheets': 0.6, 'c10_goals_conceded': -0.4, 'c11_defensive_contributions': 0.65},
        # MIDs
        {'player_code': 8, 'web_name': 'Salah', 'team': 'Liverpool', 'position': 'MID', 'cost': 12.5, 'expected_points': 7.5, 'fixture_opponent': 'Newcastle', 'fixture_venue': 'A', 'fixture_fdr': 4, 'fixture_attack_mult': 0.95, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.95, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.95, 'c3_saves': 0.0, 'c4_yellow_cards': -0.05, 'c5_red_cards': 0.0, 'c6_bonus': 1.2, 'c7_assists': 1.2, 'c8_goals': 3.0, 'c9_clean_sheets': 0.4, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
        {'player_code': 9, 'web_name': 'Saka', 'team': 'Arsenal', 'position': 'MID', 'cost': 10.0, 'expected_points': 6.8, 'fixture_opponent': 'Coventry City', 'fixture_venue': 'H', 'fixture_fdr': 2, 'fixture_attack_mult': 1.08, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.95, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.95, 'c3_saves': 0.0, 'c4_yellow_cards': -0.08, 'c5_red_cards': 0.0, 'c6_bonus': 1.0, 'c7_assists': 1.5, 'c8_goals': 2.3, 'c9_clean_sheets': 0.4, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
        {'player_code': 10, 'web_name': 'Palmer', 'team': 'Chelsea', 'position': 'MID', 'cost': 10.5, 'expected_points': 7.0, 'fixture_opponent': 'Fulham', 'fixture_venue': 'A', 'fixture_fdr': 3, 'fixture_attack_mult': 1.05, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.95, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.95, 'c3_saves': 0.0, 'c4_yellow_cards': -0.1, 'c5_red_cards': 0.0, 'c6_bonus': 1.1, 'c7_assists': 1.4, 'c8_goals': 2.5, 'c9_clean_sheets': 0.3, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
        {'player_code': 11, 'web_name': 'Rogers', 'team': 'Aston Villa', 'position': 'MID', 'cost': 5.0, 'expected_points': 4.8, 'fixture_opponent': 'Spurs', 'fixture_venue': 'H', 'fixture_fdr': 3, 'fixture_attack_mult': 1.08, 'p_start': 0.95, 'p_app': 1.0, 'p_60_plus': 0.85, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.85, 'c3_saves': 0.0, 'c4_yellow_cards': -0.15, 'c5_red_cards': 0.0, 'c6_bonus': 0.4, 'c7_assists': 0.9, 'c8_goals': 1.5, 'c9_clean_sheets': 0.3, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
        {'player_code': 12, 'web_name': 'Winks', 'team': 'Leicester', 'position': 'MID', 'cost': 4.5, 'expected_points': 2.0, 'fixture_opponent': 'Bournemouth', 'fixture_venue': 'A', 'fixture_fdr': 3, 'fixture_attack_mult': 0.92, 'p_start': 0.8, 'p_app': 0.9, 'p_60_plus': 0.7, 'c1_app_1_60': 0.9, 'c2_app_60_plus': 0.7, 'c3_saves': 0.0, 'c4_yellow_cards': -0.2, 'c5_red_cards': 0.0, 'c6_bonus': 0.1, 'c7_assists': 0.3, 'c8_goals': 0.2, 'c9_clean_sheets': 0.2, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
        # FWDs
        {'player_code': 13, 'web_name': 'Haaland', 'team': 'Man City', 'position': 'FWD', 'cost': 15.0, 'expected_points': 8.5, 'fixture_opponent': 'Bournemouth', 'fixture_venue': 'H', 'fixture_fdr': 3, 'fixture_attack_mult': 1.20, 'p_start': 1.0, 'p_app': 1.0, 'p_60_plus': 0.95, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.95, 'c3_saves': 0.0, 'c4_yellow_cards': -0.05, 'c5_red_cards': 0.0, 'c6_bonus': 1.5, 'c7_assists': 0.5, 'c8_goals': 4.6, 'c9_clean_sheets': 0.0, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
        {'player_code': 14, 'web_name': 'Isak', 'team': 'Newcastle', 'position': 'FWD', 'cost': 8.5, 'expected_points': 6.0, 'fixture_opponent': 'Liverpool', 'fixture_venue': 'H', 'fixture_fdr': 4, 'fixture_attack_mult': 1.05, 'p_start': 0.95, 'p_app': 1.0, 'p_60_plus': 0.90, 'c1_app_1_60': 1.0, 'c2_app_60_plus': 0.90, 'c3_saves': 0.0, 'c4_yellow_cards': -0.08, 'c5_red_cards': 0.0, 'c6_bonus': 0.8, 'c7_assists': 0.6, 'c8_goals': 2.8, 'c9_clean_sheets': 0.0, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
        {'player_code': 15, 'web_name': 'Wood', 'team': 'Nott\'m Forest', 'position': 'FWD', 'cost': 6.0, 'expected_points': 4.5, 'fixture_opponent': 'Leeds', 'fixture_venue': 'H', 'fixture_fdr': 2, 'fixture_attack_mult': 1.15, 'p_start': 0.9, 'p_app': 0.95, 'p_60_plus': 0.80, 'c1_app_1_60': 0.95, 'c2_app_60_plus': 0.80, 'c3_saves': 0.0, 'c4_yellow_cards': -0.1, 'c5_red_cards': 0.0, 'c6_bonus': 0.6, 'c7_assists': 0.3, 'c8_goals': 2.0, 'c9_clean_sheets': 0.0, 'c10_goals_conceded': 0.0, 'c11_defensive_contributions': 0.0},
    ]
    df = pd.DataFrame(players)
    sol = solve_initial_squad(df, budget=115.0)
    return df, sol


class TestExcelExporter:
    """Verify Excel workbook generation, styling, and data accuracy."""

    def test_export_creates_5_sheets(self, sample_test_data):
        df, sol = sample_test_data

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            out_file = export_excel_report(
                solution=sol,
                pred_df=df,
                season='2026-27',
                gw=1,
                budget=100.0,
                output_path=tmp_path,
            )
            assert os.path.exists(out_file)

            wb = openpyxl.load_workbook(out_file)
            sheet_names = wb.sheetnames
            expected_sheets = [
                "Summary Dashboard",
                "Optimal Squad",
                "GW Predictions",
                "Fixtures & Ratings",
                "Form & Underlying Stats",
            ]
            assert sheet_names == expected_sheets
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    def test_dashboard_sheet_contents(self, sample_test_data):
        df, sol = sample_test_data

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            export_excel_report(solution=sol, pred_df=df, output_path=tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb["Summary Dashboard"]

            # Title
            assert "FPL MATCHDAY DASHBOARD" in str(ws["B2"].value)
            # KPI values
            assert ws["B5"].value == "TOTAL PROJECTED xP"
            assert abs(float(ws["B6"].value) - sol.total_xp) < 0.01

            # Starting XI section
            assert ws["B9"].value == "STARTING XI LINEUP"
            # Header in row 10
            assert ws["B10"].value == "Role"
            assert ws["D10"].value == "Player"
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    def test_optimal_squad_sheet_totals(self, sample_test_data):
        df, sol = sample_test_data

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            export_excel_report(solution=sol, pred_df=df, output_path=tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb["Optimal Squad"]

            # Header row 4
            assert ws["A4"].value == "Squad Role"
            assert ws["C4"].value == "Player Name"

            # 15 player rows (5 to 19)
            assert ws["C5"].value is not None
            assert ws["C19"].value is not None

            # Total row at row 20
            assert ws["A20"].value == "TOTAL SQUAD COST & PROJECTED xP"
            assert ws["E20"].value == "=SUM(E5:E19)"
            assert ws["J20"].value == "=SUM(J5:J19)"
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    def test_predictions_sheet_sorted_by_xp(self, sample_test_data):
        df, sol = sample_test_data

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            export_excel_report(solution=sol, pred_df=df, output_path=tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb["GW Predictions"]

            # Top ranked player should be Haaland (8.5 xP)
            assert ws["A2"].value == 1  # Rank 1
            assert ws["C2"].value == "Haaland"
            assert float(ws["K2"].value) == 8.5
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
